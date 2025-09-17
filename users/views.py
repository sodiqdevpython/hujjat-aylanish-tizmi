from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.http import HttpResponseForbidden
from . import forms
from .choices import Role, DocumentStatus
from .models import User, Faculty, Department, Document, DocumentType, MainWorkPlan, AddRequirement, PlanResponse, \
    SendRequest, Notification, AcademicYear
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from datetime import datetime
from calendar import month_name
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.views.decorators.http import require_POST
from collections import defaultdict
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from decimal import Decimal
from collections import defaultdict
from django.urls import reverse


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = forms.LoginForm(request.POST or None)
    error_message = None

    if request.method == 'POST':
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                error_message = "Foydalanuvchi nomi yoki parol noto‘g‘ri"

    return render(request, 'login.html', {
        'form': form,
        'error_message': error_message
    })


class TeacherListView(LoginRequiredMixin, ListView):
    model = User
    template_name = "read/teacher_list.html"  # ↓ 2-qismda shu nomdagi fayl
    context_object_name = "page_obj"
    paginate_by = 20

    def get_queryset(self):
        qs = (
            User.objects.select_related("department__faculty")
            .filter(role__in=[Role.OQITUVCHI, Role.MUDIR, Role.PROREKTOR,
                              Role.DEKAN, Role.SUPERADMIN])  # xohlagan rollarni qoldiring
        )

        # Foydalanuvchi turi (role) bo‘yicha filtr
        role = self.request.GET.get("user_type")
        if role:
            qs = qs.filter(role=role)

        # Fakultet bo‘yicha filtr
        fac_id = self.request.GET.get("faculty_id")
        if fac_id:
            qs = qs.filter(department__faculty_id=fac_id)

        # Kafedra bo‘yicha filtr
        dep_id = self.request.GET.get("department_id")
        if dep_id:
            qs = qs.filter(department_id=dep_id)

        return qs.order_by("last_name", "first_name")

    # Sahifaga qo‘shimcha kontekst uzatamiz
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "user_types": Role.choices,
            "faculties": Faculty.objects.all(),
            "departments": Department.objects.all(),
            "selected_user_type": self.request.GET.get("user_type", ""),
            "selected_faculty_id": self.request.GET.get("faculty_id", ""),
            "selected_dept_id": self.request.GET.get("department_id", ""),
        })
        return ctx

    def dispatch(self, request, *args, **kwargs):
        if request.user.role == Role.OQITUVCHI:
            return HttpResponseForbidden("O'qituvchiga bu sahifaga kirish mumkin emas")
        return super().dispatch(request, *args, **kwargs)


@login_required
def create_teacher(request):
    if request.user.role != Role.MUDIR or not request.user.department:
        return HttpResponseForbidden("Faqat kafedra mudiri va kafedrasi mavjud foydalanuvchi qo‘shishi mumkin.")

    if request.method == 'POST':
        form = forms.TeacherCreateForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            new_user = form.save(commit=False)

            # Majburan kafedrani o‘zlashtiramiz
            new_user.department = request.user.department
            new_user.save()
            return redirect('teacher_list')
    else:
        form = forms.TeacherCreateForm(user=request.user)

    return render(request, 'create/teacher_create.html', {'form': form})


@login_required
def user_profile_detail(request, pk):
    user_profile = get_object_or_404(User, pk=pk)

    # Yuklangan hujjatlar
    documents_qs = (
        Document.objects
        .filter(upload_user=user_profile)
        .select_related("document_type")
    )

    # Hujjat turi bo‘yicha filtr
    doc_type_id = request.GET.get("doc_type")
    if doc_type_id:
        documents_qs = documents_qs.filter(document_type_id=doc_type_id)

    document_types = DocumentType.objects.all()

    # Oxirgi 6 oylik sana ro‘yxati (tuple -> (yil, oy))
    today = datetime.today()
    last_6_months = [(today.year if today.month - i > 0 else today.year - 1,
                      (today.month - i - 1) % 12 + 1) for i in range(6)]

    # Faollik statistikasi (created → NOT created_at)
    faollik_raw = (
        documents_qs
        .filter(created__year__in=[y for y, m in last_6_months],
                created__month__in=[m for y, m in last_6_months])
        .values("created__year", "created__month")
        .annotate(count=Count("id"))
    )

    # Chart uchun etiketkalar va sonlar
    month_labels = [f"{month_name[m]} {y}" for y, m in reversed(last_6_months)]
    month_counts = []
    for y, m in reversed(last_6_months):
        count = next((item["count"] for item in faollik_raw
                      if item["created__year"] == y and item["created__month"] == m), 0)
        month_counts.append(count)

    # Ruxsat: kafedra mudiri va shu kafedradan bo‘lsa
    has_permission_for_update_or_adding_work = (
            request.user.role == Role.MUDIR and
            request.user.department == user_profile.department
    )

    context = {
        "user_profile": user_profile,
        "documents": documents_qs,
        "document_types": document_types,
        "selected_type": doc_type_id,
        "faollik_months": month_labels,
        "faollik_counts": month_counts,
        "has_permission_for_update_or_adding_work": has_permission_for_update_or_adding_work
    }

    return render(request, 'read/user_profile_detail.html', context)


@login_required
def teacher_update_view(request, pk):
    user = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        form = forms.TeacherUpdateForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Foydalanuvchi ma'lumotlari yangilandi.")
            return redirect('user_profile_detail', pk=user.pk)
    else:
        form = forms.TeacherUpdateForm(instance=user)

    return render(request, 'update/teacher_update.html', {'form': form})


@login_required
def ilmiy_ishlar_list(request):
    query = request.GET.get("q", "")
    doc_type = request.GET.get("doc_type", "")
    faculty_id = request.GET.get("faculty", "")
    department_id = request.GET.get("department", "")

    # Barcha ilmiy ishlar
    documents = Document.objects.select_related(
        'upload_user', 'document_type', 'upload_user__department__faculty'
    )

    # Qidiruv
    if query:
        documents = documents.filter(
            Q(title__icontains=query) |
            Q(short_description__icontains=query) |
            Q(upload_user__first_name__icontains=query) |
            Q(upload_user__last_name__icontains=query)
        )

    # Hujjat turi bo‘yicha filtr
    if doc_type:
        documents = documents.filter(document_type_id=doc_type)

    # Fakultet bo‘yicha filtr
    if faculty_id:
        documents = documents.filter(
            upload_user__department__faculty_id=faculty_id
        )

    # Kafedra bo‘yicha filtr
    if department_id:
        documents = documents.filter(
            upload_user__department_id=department_id
        )

    # Paginatsiya (9 kartochka / sahifa)
    paginator = Paginator(documents.order_by("-created"), 9)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "query": query,
        "doc_types": DocumentType.objects.all(),
        "selected_doc_type": doc_type,
        "faculties": Faculty.objects.all(),
        "departments": Department.objects.all(),
        "selected_faculty": faculty_id,
        "selected_department": department_id,
    }
    return render(request, "read/ilmiy_ishlar_list.html", context)


@login_required
def ilmiy_ish_detail(request, pk):
    document = get_object_or_404(Document, pk=pk)

    # upload_user yuklagan boshqa 4 ta ish
    related_documents = (
        Document.objects
        .filter(upload_user=document.upload_user)
        .exclude(id=document.id)
        .order_by('-created')[:4]
    )

    return render(request, "read/document_detail.html", {
        "document": document,
        "related_documents": related_documents,
        "document_status_choices": DocumentStatus.choices,
    })


@require_POST
@login_required
def update_document_status(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.user.role != Role.MUDIR or request.user.department != doc.upload_user.department:
        return redirect('ilmiy_ish_detail', pk=pk)

    new_status = request.POST.get("status")
    if new_status in dict(DocumentStatus.choices):
        doc.status = new_status
        doc.is_confirmed = (new_status == DocumentStatus.APPROVED)
        doc.save()

    return redirect('ilmiy_ish_detail', pk=pk)


@login_required
def add_requirement(request, teacher_id):
    teacher = get_object_or_404(User, pk=teacher_id)
    # faqat o‘sha teacher kafedrasi mudiri qo‘shishi mumkin
    if request.user.role != Role.MUDIR or request.user.department != teacher.department:
        return HttpResponseForbidden("Siz bu foydalanuvchiga ish reja qo‘sha olmaysiz.")

    if request.method == 'POST':
        form = forms.AddRequirementForm(request.POST, teacher=teacher)
        if form.is_valid():
            form.save()
            return redirect('user_profile_detail', pk=teacher.pk)
    else:
        form = forms.AddRequirementForm(teacher=teacher)

    return render(request, 'create/add_requirement.html', {
        'form': form,
        'teacher': teacher
    })


@login_required
def work_plan_report(request):
    if request.user.role != Role.MUDIR:
        return HttpResponseForbidden()

    # ---- Filterlar ----
    fac_id = request.GET.get("faculty")
    dept_id = request.GET.get("department")
    year_id = request.GET.get("year")  # ⬅️ yangi qo‘shildi

    faculties = Faculty.objects.all()
    departments = Department.objects.filter(faculty_id=fac_id) if fac_id else Department.objects.none()
    years = AcademicYear.objects.all()

    teachers = User.objects.filter(role=Role.OQITUVCHI)
    if fac_id:
        teachers = teachers.filter(department__faculty_id=fac_id)
    if dept_id:
        teachers = teachers.filter(department_id=dept_id)
    teachers = teachers.order_by("last_name", "first_name")

    # ---- Jadval ustunlari (unchalik o‘zgarmaydi) ----
    main_plans = MainWorkPlan.objects.prefetch_related("subwork")
    columns, main_meta = [], []
    for mp in main_plans:
        subs = list(mp.subwork.all()) or [None]
        columns.extend([(mp, sp) for sp in subs])
        main_meta.append({"mp": mp, "subs": subs, "colspan": len(subs) * 2})

    # ---- Statistik ma’lumot ----
    stat = defaultdict(lambda: {"planned": Decimal("0.0"), "actual": Decimal("0.0")})

    # Reja miqdori
    req_qs = AddRequirement.objects.filter(teacher__in=teachers)
    if year_id:
        req_qs = req_qs.filter(academic_year_id=year_id)
    for req in req_qs:
        main_id = req.main_plan_id or (req.sub_plan.parent_id if req.sub_plan else None)
        sub_id = req.sub_plan_id
        key = (req.teacher_id, main_id, sub_id)
        stat[key]["planned"] += req.quantity_planned

    # Amalda bajarilgan
    pr_qs = PlanResponse.objects.select_related("requirement").filter(
        requirement__teacher__in=teachers
    )
    if year_id:
        pr_qs = pr_qs.filter(academic_year_id=year_id)
    for pr in pr_qs:
        req = pr.requirement
        main_id = req.main_plan_id or (req.sub_plan.parent_id if req.sub_plan else None)
        sub_id = req.sub_plan_id
        key = (req.teacher_id, main_id, sub_id)
        stat[key]["actual"] += pr.quantity_actual

    # ---- Jadval satrlari ----
    table_rows = []
    for idx, t in enumerate(teachers, 1):
        row_cells = []
        for mp, sp in columns:
            cell = stat[(t.id, mp.id, sp.id if sp else None)]
            row_cells.append(cell)
        table_rows.append({"idx": idx, "teacher": t, "cells": row_cells})

    paginator = Paginator(table_rows, 100)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ---- Footer ----
    footer = []
    for mp, sp in columns:
        p_sum = sum(stat[(t.id, mp.id, sp.id if sp else None)]["planned"] for t in teachers)
        a_sum = sum(stat[(t.id, mp.id, sp.id if sp else None)]["actual"] for t in teachers)
        footer.append({"planned": p_sum, "actual": a_sum})

    return render(
        request,
        "read/work_plan_report.html",
        {
            "faculties": faculties,
            "departments": departments,
            "years": years,  # ⬅️ Templatega yuboriladi
            "selected_faculty": fac_id or "",
            "selected_department": dept_id or "",
            "selected_year": year_id or "",
            "main_meta": main_meta,
            "columns": columns,
            "table_rows": table_rows,
            "footer": footer,
            "page_obj": page_obj,
        },
    )


@login_required
def work_plan_export(request):
    if request.user.role != Role.MUDIR:
        return HttpResponseForbidden()

    # === 1. Filtrlar ===
    fac_id = request.GET.get("faculty")
    dept_id = request.GET.get("department")
    year_id = request.GET.get("academic_year")  # 🔥 academic year filter

    teachers = User.objects.filter(role=Role.OQITUVCHI)
    if fac_id:
        teachers = teachers.filter(department__faculty_id=fac_id)
    if dept_id:
        teachers = teachers.filter(department_id=dept_id)
    teachers = teachers.order_by("last_name", "first_name")

    # === 2. Ustunlar tartibi (faqat shu yil) ===
    main_plans = MainWorkPlan.objects.prefetch_related("subwork")
    if year_id:
        main_plans = main_plans.filter(academic_year_id=year_id)  # 🔥 yil bo‘yicha filter

    columns = []
    for mp in main_plans:
        subs = list(mp.subwork.all()) or [None]  # child yo‘q bo‘lsa bitta bo‘sh sub
        columns.extend([(mp, sp) for sp in subs])

    # === 3. Reja / Amalda yig‘indilar ===
    stat = defaultdict(lambda: {"planned": Decimal("0"), "actual": Decimal("0")})

    reqs = AddRequirement.objects.filter(teacher__in=teachers)
    if year_id:
        reqs = reqs.filter(academic_year_id=year_id)  # 🔥 yil bo‘yicha filter

    for req in reqs:
        mp = req.main_plan or (req.sub_plan.parent if req.sub_plan else None)
        if not mp:
            continue
        key = (req.teacher_id, mp.id, req.sub_plan_id)
        stat[key]["planned"] += req.quantity_planned

    prs = PlanResponse.objects.select_related("requirement").filter(
        requirement__teacher__in=teachers
    )
    if year_id:
        prs = prs.filter(requirement__academic_year_id=year_id)  # 🔥 yil bo‘yicha filter

    for pr in prs:
        req = pr.requirement
        mp = req.main_plan or (req.sub_plan.parent if req.sub_plan else None)
        if not mp:
            continue
        key = (req.teacher_id, mp.id, req.sub_plan_id)
        stat[key]["actual"] += pr.quantity_actual

    # === 4. Excel yaratish ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ish reja"

    blue = PatternFill("solid", fgColor="BDD7EE")  # katta sarlavha
    gray = PatternFill("solid", fgColor="D9D9D9")  # kichik sarlavha
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrapText=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    # === 4-a 3-qavatli sarlavha ===
    row = 1
    col = 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col)
    ws.cell(row, col, "#").fill = blue
    ws.cell(row, col).font = bold
    ws.cell(row, col).alignment = center
    col += 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col)
    ws.cell(row, col, "F.I.Sh").fill = blue
    ws.cell(row, col).font = bold
    ws.cell(row, col).alignment = center
    col += 1

    # 1-qavat: MainWorkPlan nomi
    for mp in main_plans:
        span = (len(mp.subwork.all()) or 1) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        cell = ws.cell(row, col, mp.name)
        cell.fill = blue
        cell.font = bold
        cell.alignment = center
        col += span
    row += 1
    col = 3

    # 2-qavat: SubWorkPlan nomi
    for mp in main_plans:
        subs = list(mp.subwork.all()) or [None]
        for sp in subs:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            name = sp.name if sp else ""
            cell = ws.cell(row, col, name)
            cell.fill = gray
            cell.font = bold
            cell.alignment = center
            col += 2
    row += 1
    col = 3

    # 3-qavat: Reja/Amalda
    for mp, sp in columns:
        ws.cell(row, col, "Reja").fill = gray
        ws.cell(row, col).font = bold
        ws.cell(row, col).alignment = center
        ws.cell(row, col + 1, "Amalda").fill = gray
        ws.cell(row, col + 1).font = bold
        ws.cell(row, col + 1).alignment = center
        col += 2

    # === 4-b  O‘qituvchi satrlari ===
    excel_row = 4
    idx = 1
    for t in teachers:
        ws.cell(excel_row, 1, idx)
        ws.cell(excel_row, 2, f"{t.last_name} {t.first_name}")
        col = 3
        for mp, sp in columns:
            key = (t.id, mp.id, sp.id if sp else None)
            dat = stat[key]
            ws.cell(excel_row, col, float(dat["planned"])).alignment = center
            ws.cell(excel_row, col + 1, float(dat["actual"])).alignment = center
            col += 2
        idx += 1
        excel_row += 1

    # === 4-c  Jami qatori ===
    ws.cell(excel_row, 1, "Jami")
    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=2)
    ws.cell(excel_row, 1).font = bold
    ws.cell(excel_row, 1).alignment = center
    col = 3
    for mp, sp in columns:
        planned_sum = sum(stat[(t.id, mp.id, sp.id if sp else None)]["planned"] for t in teachers)
        actual_sum = sum(stat[(t.id, mp.id, sp.id if sp else None)]["actual"] for t in teachers)
        ws.cell(excel_row, col, float(planned_sum)).font = bold
        ws.cell(excel_row, col).alignment = center
        ws.cell(excel_row, col + 1, float(actual_sum)).font = bold
        ws.cell(excel_row, col + 1).alignment = center
        col += 2

    # === 4-d  Chiziqlar va ustun eni ===
    for r in ws.iter_rows(min_row=1, max_row=excel_row, min_col=1, max_col=col - 1):
        for c in r:
            c.border = thin
    for c in range(1, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12

    # === 5. HTTP javob ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"work_plan_report_{year_id or 'all'}.xlsx"  # 🔥 fayl nomida yil
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def doc_approve_detail(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.user.role != Role.MUDIR or request.user.department != doc.upload_user.department:
        return HttpResponseForbidden()

    if request.method == "POST":
        form = forms.ApproveForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data["action"]  # approve / reject
            comment = form.cleaned_data["comment"]
            if action == "approve":
                doc.is_confirmed = True
                doc.status = DocumentStatus.APPROVED
                doc.save()

                # reja → amalda
                req = AddRequirement.objects.filter(
                    teacher=doc.upload_user, sub_plan__isnull=False).last()
                if req:
                    PlanResponse.objects.create(
                        requirement=req, document=doc, quantity_actual=1)

            else:  # reject
                SendRequest.objects.create(
                    document=doc,
                    requirement=None,
                    status=DocumentStatus.REJECTED,
                    rejected_message=comment,
                    who_rejected=request.user)

                doc.status = DocumentStatus.REJECTED
                doc.save()
            # bildirshnoma: o‘qituvchiga
            Notification.objects.create(
                recipient=doc.upload_user,
                title=f"Hujjat {'tasdiqlandi' if action == 'approve' else 'rad etildi'}",
                message=comment or "-",
                url=reverse("ilmiy_ish_detail", args=[doc.id])
            )
            return redirect("work_plan_report")
    else:
        form = forms.ApproveForm()

    return render(request, "read/doc_detail.html", {"doc": doc, "form": form})


@login_required
def notification_list(request):
    # Faqat o'z xabarlari, yangi xabarlar avval
    notif_qs = Notification.objects.filter(
        recipient=request.user,
        is_read=False
    ).order_by("-created")

    # ----- Paginator (20 ta) -----
    paginator = Paginator(notif_qs, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Sahifada ko‘rinayotgan xabarlarni o‘qilgan deb belgilash
    # visible_ids = [n.id for n in page_obj if not n.is_read]
    # if visible_ids:
    #     Notification.objects.filter(id__in=visible_ids).update(is_read=True)

    return render(request, "read/notifications-list.html", {"page_obj": page_obj})


@login_required
def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.user.role == Role.OQITUVCHI:
        return redirect('teacher_dashboard')
    # ---- 1.  Ruxsat ----
    if request.user.role not in (
            Role.MUDIR,
            Role.SUPERADMIN,
            Role.PROREKTOR,
            Role.DEKAN,
    ):
        return HttpResponseForbidden()

    # ---- 2.  Asosiy raqamlar ----
    total_faculties = Faculty.objects.count()
    total_departments = Department.objects.count()
    total_users = User.objects.count()
    total_docs = Document.objects.count()
    pending_docs = Document.objects.filter(status=DocumentStatus.PENDING).count()

    # ---- 3. Fakultet → kafedra sonlari ----
    fac_breakdown = (
        Faculty.objects
        .annotate(dep_cnt=Count("department"))
        .values("title", "dep_cnt")
        .order_by("-dep_cnt")
    )

    # ---- 4. Foydalanuvchi rollari bo‘yicha taqsimot ----
    role_stats = (
        User.objects
        .values("role")
        .annotate(total=Count("id"))
    )
    # Turn into dict {display: count}
    role_counts = {
        dict(Role.choices)[item["role"]]: item["total"]
        for item in role_stats
    }

    # ---- 5. Hujjatlar holati (tasdiqlangan/pending/rejected) ----
    doc_status_stats = (
        Document.objects
        .values("status")
        .annotate(total=Count("id"))
    )
    status_counts = {
        dict(DocumentStatus.choices)[s["status"]]: s["total"]
        for s in doc_status_stats
    }

    # ---- 6. Last actions (oxirgi 10 ta hujjat) ----
    recent_docs = (
        Document.objects
        .select_related("upload_user", "document_type")
        .order_by("-created")[:10]
    )

    context = {
        # blok 1
        "total_faculties": total_faculties,
        "total_departments": total_departments,
        "total_users": total_users,
        "total_docs": total_docs,
        "pending_docs": pending_docs,

        # blok 2
        "fac_breakdown": fac_breakdown,  # list of dicts
        "role_counts": role_counts,  # dict
        "status_counts": status_counts,  # dict

        # blok 3
        "recent_docs": recent_docs,
    }
    return render(request, "read/admin_dashboard.html", context)


@login_required
def teacher_dashboard(request):
    # 1) faqat O‘QITUVCHI roli
    if request.user.role != Role.OQITUVCHI:
        return HttpResponseForbidden()

    user = request.user

    # 2) Rejalar va amalda bajarilgan ishlar
    req_qs = AddRequirement.objects.filter(teacher=user)
    resp_qs = PlanResponse.objects.filter(requirement__teacher=user)

    total_req_cnt = req_qs.count()
    total_req_qty = req_qs.aggregate(q=Sum("quantity_planned"))["q"] or 0

    total_done_qty = resp_qs.aggregate(q=Sum("quantity_actual"))["q"] or 0
    progress_pct = round((total_done_qty / total_req_qty) * 100, 1) if total_req_qty else 0

    # 3) Hujjat statuslari
    status_map = dict(DocumentStatus.choices)  # {'approved':'Tasdiqlangan', ...}
    raw_status = (
        Document.objects
        .filter(upload_user=user)
        .values("status")
        .annotate(total=Count("id"))
    )
    doc_status = {name: 0 for name in status_map.values()}  # default 0
    for row in raw_status:
        doc_status[status_map[row["status"]]] = row["total"]

    # 4) Oxirgi xabarlar (o‘qilmagan Rejection + umumiy)
    unread_notifs = Notification.objects.filter(recipient=user, is_read=False)[:10]

    # 5) Oxirgi 6 ta hujjat
    recent_docs = (
        Document.objects
        .filter(upload_user=user)
        .select_related("document_type")
        .order_by("-created")[:6]
    )

    ctx = {
        "total_req_cnt": total_req_cnt,
        "total_req_qty": total_req_qty,
        "total_done_qty": total_done_qty,
        "progress_pct": progress_pct,
        "doc_status": doc_status,
        "unread_notifs": unread_notifs,
        "recent_docs": recent_docs,
    }
    return render(request, "read/teacher_dashboard.html", ctx)


@login_required
def my_profile(request):
    user = request.user

    # ----- 1.  O‘qilmagan xabarlar (eng ko‘pi 15 ta) -----
    unread_notifs = Notification.objects.filter(
        recipient=user, is_read=False
    ).order_by("-created")[:15]

    # ----- 2.  Hujjatlar ro‘yxati + filtr -----
    query = request.GET.get("q", "")
    type_id = request.GET.get("type", "")

    docs_qs = Document.objects.filter(upload_user=user).select_related("document_type")

    if query:
        docs_qs = docs_qs.filter(
            Q(title__icontains=query) |
            Q(short_description__icontains=query)
        )
    if type_id:
        docs_qs = docs_qs.filter(document_type_id=type_id)

    paginator = Paginator(docs_qs.order_by("-created"), 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "unread_notifs": unread_notifs,
        "page_obj": page_obj,
        "doc_types": DocumentType.objects.all(),
        "selected_type": type_id,
        "query": query,
    }
    return render(request, "read/my_profile.html", context)


@login_required
def my_documents(request):
    docs = Document.objects.filter(upload_user=request.user)
    query = request.GET.get("q", "")
    status = request.GET.get("status", "")

    if query:
        docs = docs.filter(
            Q(title__icontains=query) |
            Q(short_description__icontains=query)
        )
    if status:
        docs = docs.filter(status=status)

    paginator = Paginator(docs.order_by("-created"), 9)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "read/my_documents.html", {
        "page_obj": page_obj,
        "query": query,
        "selected_status": status
    })


@login_required
def my_requirements_view(request):
    teacher = request.user

    # Barcha biriktirilgan talablar va ularning javoblari
    requirements = AddRequirement.objects.filter(teacher=teacher).select_related("main_plan", "sub_plan")

    data = []
    for req in requirements:
        actual = PlanResponse.objects.filter(requirement=req).aggregate(
            total=Sum("quantity_actual")
        )["total"] or 0

        data.append({
            "plan": req.main_plan.name if req.main_plan else "-",
            "sub_plan": req.sub_plan.name if req.sub_plan else "-",
            "planned": req.quantity_planned,
            "actual": actual,
            "status": "Yakunlangan" if actual >= req.quantity_planned else "Yetmagan"
        })

    return render(request, "read/my_requirements.html", {"requirements": data})


@login_required
def my_notifications(request):
    qs = Notification.objects.filter(recipient=request.user).order_by("-created")

    paginator = Paginator(qs, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # O‘qilgan deb belgilash
    unread_ids = [n.id for n in page_obj if not n.is_read]
    if unread_ids:
        Notification.objects.filter(id__in=unread_ids).update(is_read=True)

    return render(request, "read/my_notifications.html", {"page_obj": page_obj})


@login_required
def document_create(request):
    if request.user.role != Role.OQITUVCHI:
        return HttpResponseForbidden()

    if request.method == "POST":
        form = forms.DocumentCreateForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.upload_user = request.user
            doc.status = DocumentStatus.PENDING
            doc.is_confirmed = False
            doc.save()

            # kafedra mudiriga xabar
            # mudir = request.user.department.user_set.filter(role=Role.MUDIR).first()
            # if mudir:
            #     Notification.objects.create(
            #         recipient=mudir,
            #         title="Tasdiqlash uchun yangi hujjat",
            #         message=f"{request.user.get_full_name()} «{doc.title}» yukladi",
            #         url=reverse("doc_approve_detail", args=[doc.id])
            #     )
            return redirect("my_documents")
    else:
        form = forms.DocumentCreateForm(user=request.user)

    return render(request, "create/document_create.html", {"form": form})
